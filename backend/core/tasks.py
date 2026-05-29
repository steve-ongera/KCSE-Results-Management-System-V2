"""
apps/examinations/tasks.py

Celery async tasks for the KCSE Management System.

Tasks:
  process_bulk_marks_upload  — parse CSV/Excel and bulk-create MarksEntry rows
  process_all_results        — compute grades, mean scores, and rankings
  recompute_rankings         — nightly re-rank all candidates
  send_results_notification  — SMS/email notification (stub)
"""

import io
import csv
import logging
import uuid

from celery import shared_task
from django.utils import timezone

#logger = logging.getLogger('apps.examinations')


# ─────────────────────────────────────────────────────────────────────────────
# BULK MARKS UPLOAD
# ─────────────────────────────────────────────────────────────────────────────

@shared_task(bind=True, max_retries=3, default_retry_delay=60)
def process_bulk_marks_upload(
    self,
    file_content: bytes,
    file_name: str,
    subject_paper_id: str,
    examiner_id: str,
    batch_id: str,
):
    """
    Parse a CSV or Excel file and create MarksEntry rows in bulk.

    CSV expected format (header row required):
        barcode,marks
        KNEC2024001,78
        KNEC2024002,54.5

    Excel (.xlsx/.xls) expected: same columns, first sheet.

    Creates a MarksEntry for each row.
    Rows with validation errors are collected and reported.
    Marks flagged as abnormal (0 or max) are auto-flagged.
    """

    from django.contrib.auth import get_user_model
    from .models import (
        ExaminationScript, SubjectPaper,
        MarksEntry, MarksStatus,
    )

    User = get_user_model()

    try:
        paper    = SubjectPaper.objects.get(pk=subject_paper_id)
        examiner = User.objects.get(pk=examiner_id)
    except (SubjectPaper.DoesNotExist, User.DoesNotExist) as exc:
        logger.error("Bulk upload task: reference not found — %s", exc)
        return {'status': 'error', 'message': str(exc)}

    rows      = []
    errors    = []
    created   = 0
    skipped   = 0

    # ── Parse file ───────────────────────────────────────────────────────────
    try:
        if file_name.endswith('.csv'):
            text   = file_content.decode('utf-8', errors='replace')
            reader = csv.DictReader(io.StringIO(text))
            rows   = list(reader)

        elif file_name.endswith(('.xlsx', '.xls')):
            try:
                import openpyxl
                wb   = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
                ws   = wb.active
                hdrs = [str(c.value).strip().lower() for c in next(ws.iter_rows())]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(hdrs, row)))
            except ImportError:
                logger.warning("openpyxl not installed — install it for Excel support")
                return {'status': 'error', 'message': 'openpyxl not installed for Excel support.'}

        else:
            return {'status': 'error', 'message': f'Unsupported file type: {file_name}'}

    except Exception as exc:
        logger.error("Bulk upload parse error: %s", exc)
        raise self.retry(exc=exc)

    # ── Process rows ─────────────────────────────────────────────────────────
    entries_to_create = []

    for i, row in enumerate(rows, start=2):
        barcode = str(row.get('barcode', '')).strip()
        raw_marks = row.get('marks', '')

        if not barcode:
            errors.append({'row': i, 'error': 'Missing barcode'})
            skipped += 1
            continue

        try:
            marks = float(str(raw_marks).strip())
        except (ValueError, TypeError):
            errors.append({'row': i, 'barcode': barcode, 'error': f'Invalid marks: {raw_marks}'})
            skipped += 1
            continue

        if marks < 0 or marks > paper.max_marks:
            errors.append({
                'row': i, 'barcode': barcode,
                'error': f'Marks {marks} out of range (0–{paper.max_marks})'
            })
            skipped += 1
            continue

        try:
            script = ExaminationScript.objects.get(barcode=barcode, subject_paper=paper)
        except ExaminationScript.DoesNotExist:
            errors.append({'row': i, 'barcode': barcode, 'error': 'Script not found'})
            skipped += 1
            continue

        # Skip if already entered (prevent duplicate bulk uploads)
        if MarksEntry.objects.filter(candidate=script.candidate, subject_paper=paper).exists():
            errors.append({'row': i, 'barcode': barcode, 'error': 'Marks already entered'})
            skipped += 1
            continue

        is_abnormal = (marks == 0 or marks == paper.max_marks)
        entries_to_create.append(
            MarksEntry(
                script=script,
                candidate=script.candidate,
                subject_paper=paper,
                examiner=examiner,
                marks=marks,
                status=MarksStatus.ENTERED,
                is_abnormal=is_abnormal,
                abnormality_note=(
                    'Bulk upload: zero marks.' if marks == 0
                    else 'Bulk upload: perfect score.' if is_abnormal
                    else ''
                ),
                is_bulk_uploaded=True,
                upload_batch_id=batch_id,
            )
        )

    # ── Bulk create ──────────────────────────────────────────────────────────
    if entries_to_create:
        MarksEntry.objects.bulk_create(entries_to_create, batch_size=500,
                                       ignore_conflicts=True)
        created = len(entries_to_create)

    result = {
        'status':   'completed',
        'batch_id': batch_id,
        'total':    len(rows),
        'created':  created,
        'skipped':  skipped,
        'errors':   errors[:50],  # Cap error list to avoid enormous task result
    }
    logger.info("Bulk upload %s: %d created, %d skipped", batch_id, created, skipped)
    return result


# ─────────────────────────────────────────────────────────────────────────────
# RESULTS PROCESSING
# ─────────────────────────────────────────────────────────────────────────────

@shared_task(bind=True, soft_time_limit=25 * 60, time_limit=30 * 60)
def process_all_results(self, examination_year_id: str, user_id: str):
    """
    Compute grades, mean scores, and rankings for every approved candidate
    in the given examination year.

    Steps:
      1. For each approved candidate, compute SubjectResult rows from MarksEntry.
      2. Compute CandidateResult (mean grade, mean points).
      3. Assign school, county, and national rankings.

    This task may run for several minutes on large datasets.
    Progress is logged to the task logger.
    """

    from django.contrib.auth import get_user_model
    from .models import (
        ExaminationYear, Candidate, SubjectResult,
        MarksEntry, MarksStatus, RegistrationStatus,
        AuditLog, AuditAction,
    )
    from .grading import marks_to_grade, compute_candidate_result, _compute_rankings

    User = get_user_model()

    try:
        year = ExaminationYear.objects.get(pk=examination_year_id)
        user = User.objects.get(pk=user_id)
    except (ExaminationYear.DoesNotExist, User.DoesNotExist) as exc:
        logger.error("process_all_results: %s", exc)
        return {'status': 'error', 'message': str(exc)}

    logger.info("Starting results processing for KCSE %s", year.year)

    candidates = Candidate.objects.filter(
        examination_year=year,
        registration_status=RegistrationStatus.KNEC_APPR,
    ).select_related('examination_center')

    processed = skipped = errors = 0
    subject_results_to_create = []

    for candidate in candidates:
        try:
            # Get all locked marks for this candidate in this year
            marks_qs = MarksEntry.objects.filter(
                candidate=candidate,
                subject_paper__subject__candidatesubject__candidate=candidate,
                status=MarksStatus.LOCKED,
            ).select_related('subject_paper__subject')

            # Group by subject — sum paper marks
            subject_marks = {}
            for entry in marks_qs:
                subj = entry.subject_paper.subject
                paper_num = entry.subject_paper.paper_number
                if subj.pk not in subject_marks:
                    subject_marks[subj.pk] = {'subject': subj, 'papers': {}}
                subject_marks[subj.pk]['papers'][f'paper_{paper_num}'] = float(entry.marks)

            if not subject_marks:
                skipped += 1
                continue

            # Create/update SubjectResult for each subject
            for subj_id, data in subject_marks.items():
                subj      = data['subject']
                papers    = data['papers']
                total_raw = sum(papers.values())

                # Scale to 100 if subject max_marks differs
                scale = 100 / subj.max_marks if subj.max_marks else 1
                total = min(total_raw * scale, 100)

                points, grade = marks_to_grade(total)

                subject_results_to_create.append(
                    SubjectResult(
                        candidate=candidate,
                        subject=subj,
                        examination_year=year,
                        paper_marks=papers,
                        total_marks=round(total_raw, 2),
                        moderation_adjustment=0,
                        moderated_marks=round(total, 2),
                        grade=grade,
                        points=points,
                    )
                )

            processed += 1

        except Exception as exc:
            logger.error("Error processing candidate %s: %s", candidate.index_number, exc)
            errors += 1

    # Bulk create subject results (delete old ones first for idempotency)
    from .models import SubjectResult as SR
    SR.objects.filter(examination_year=year).delete()
    SR.objects.bulk_create(subject_results_to_create, batch_size=500)

    # Compute aggregate (CandidateResult) for each candidate
    for candidate in candidates:
        try:
            compute_candidate_result(candidate, year, processed_by=user)
        except Exception as exc:
            logger.error("CandidateResult error for %s: %s", candidate.index_number, exc)

    # Assign rankings
    _compute_rankings(year)

    AuditLog.log(
        action=AuditAction.UPDATE,
        user=user,
        description=(
            f"Results processed for KCSE {year.year}: "
            f"{processed} candidates, {skipped} skipped, {errors} errors."
        ),
        object_type='ExaminationYear',
        object_id=str(year.pk),
    )

    logger.info(
        "Results processing complete: %d processed, %d skipped, %d errors",
        processed, skipped, errors,
    )
    return {
        'status':    'completed',
        'year':      year.year,
        'processed': processed,
        'skipped':   skipped,
        'errors':    errors,
    }


# ─────────────────────────────────────────────────────────────────────────────
# NIGHTLY RANKINGS RECOMPUTATION
# ─────────────────────────────────────────────────────────────────────────────

@shared_task
def recompute_rankings():
    """
    Nightly Celery Beat task.
    Re-runs ranking computation for all published years to keep
    rankings accurate even if results were withheld/updated.
    """

    from .models import ExaminationYear
    from .grading import _compute_rankings

    published_years = ExaminationYear.objects.filter(results_published=True)
    for year in published_years:
        try:
            _compute_rankings(year)
            logger.info("Rankings recomputed for KCSE %s", year.year)
        except Exception as exc:
            logger.error("Ranking error for KCSE %s: %s", year.year, exc)

    return {'recomputed': [y.year for y in published_years]}


# ─────────────────────────────────────────────────────────────────────────────
# RESULTS NOTIFICATION (Stub — integrate with Africa's Talking or Twilio)
# ─────────────────────────────────────────────────────────────────────────────

@shared_task
def send_results_notification(candidate_id: str, channel: str = 'sms'):
    """
    Send a results-available notification to a candidate via SMS or email.

    channel: 'sms' | 'email'

    TODO: Integrate with:
      - Africa's Talking SMS API for Kenya (https://africastalking.com)
      - Django's send_mail for email
    """

    from .models import Candidate, CandidateResult

    try:
        candidate = Candidate.objects.select_related(
            'examination_center', 'examination_year', 'result'
        ).get(pk=candidate_id)
    except Candidate.DoesNotExist:
        logger.error("Notification: candidate %s not found", candidate_id)
        return

    try:
        result = candidate.result
    except CandidateResult.DoesNotExist:
        logger.warning("Notification: no result for candidate %s", candidate_id)
        return

    message = (
        f"Dear {candidate.full_name.split()[0].title()}, "
        f"your KCSE {candidate.examination_year.year} results are now available. "
        f"Mean Grade: {result.mean_grade}. "
        f"Visit results.knec.ac.ke to view full results. — KNEC"
    )

    if channel == 'sms':
        # TODO: replace with Africa's Talking SDK call
        logger.info("[SMS STUB] To: %s | Message: %s", candidate.index_number, message)
    elif channel == 'email':
        # TODO: replace with send_mail call
        logger.info("[EMAIL STUB] To: %s | Message: %s", candidate.index_number, message)